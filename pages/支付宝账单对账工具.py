import streamlit as st
import pandas as pd
import json
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill


# ======================================
#  对账逻辑
# ======================================
def generate_reconcile_table(df, tax_area, owner, period):
    # 1. 清洗企业名称（对方账号去掉邮箱）
    df["企业名称"] = (
        df["对方账号"]
        .str.replace(r"\(.*", "", regex=True)  # 去掉括号后内容
        .str.replace(r"@.*", "", regex=True)  # 去掉邮箱后内容
        .str.strip()  # 去掉收尾空格
    )

    # 2. 计算实发金额
    df["收入金额（+元）"] = pd.to_numeric(df["收入金额（+元）"], errors="coerce").fillna(0)
    df["支出金额（-元）"] = pd.to_numeric(df["支出金额（-元）"], errors="coerce").fillna(0)

    grouped = df.groupby("企业名称").agg({
        "收入金额（+元）": "sum",
        "支出金额（-元）": "sum"
    }).reset_index()

    grouped["支付宝实发金额（已付）"] = (grouped["收入金额（+元）"] - grouped["支出金额（-元）"]).round(2)

    with open("config/rate_map.json", "r", encoding="utf-8") as f:
        rate_map = json.load(f)
    grouped["服务费率"] = grouped["企业名称"].map(rate_map).fillna(0)

    # 支付宝服务费
    grouped["支付宝服务费（未付）"] = (grouped["支付宝实发金额（已付）"] * grouped["服务费率"] / 100).round(2)

    # 其它字段（置空或计算）
    grouped["税收地"] = tax_area
    grouped["对账周期"] = period
    grouped["负责人"] = owner

    grouped["合计金额（开票金额）"] = (grouped["支付宝实发金额（已付）"] + grouped["支付宝服务费（未付）"]).round(2)
    grouped["开票信息（甲方提供过）"] = "公司名称:" + grouped["企业名称"]

    grouped["开票时间"] = ""
    grouped["开票金额"] = grouped["合计金额（开票金额）"].round(2)
    grouped["应付服务费"] = grouped["支付宝服务费（未付）"].round(2)
    grouped["是否付服务费"] = ""
    grouped["实际应付服务费（按照6.72收取）"] = ""
    grouped["返税金额：6.72-实际费率"] = ""
    # 序号
    # 序号
    grouped.insert(0, "序号", range(1, len(grouped) + 1))

    # 财务最终列顺序
    final_columns = [
        "序号",
        "税收地",
        "企业名称",
        "服务费率",
        "费率",
        "对账周期",
        "负责人",
        "支付宝实发金额（已付）",
        "支付宝服务费（未付）",
        "网商银行实发金额",
        "网商银行服务费（未付）",
        "合计金额（开票金额）",
        "开票信息（甲方提供过）",
        "开票时间",
        "开票金额",
        "应付服务费",
        "是否付服务费",
        "实际应付服务费（按照6.72收取）",
        "返税金额：6.72-实际费率",
    ]
    grouped = grouped.reindex(columns=final_columns)
    return grouped


# ======================================
#                Web UI
# ======================================
st.title("支付宝账单对账工具")

uploaded = st.file_uploader("上传支付宝账单 Excel 文件（必须是 .xlsx）", type=["xlsx"])

st.write("---")

tax_area = st.text_input("税收地")
owner = st.text_input("负责人")
period = st.text_input("对账周期（例如：6.1-6.19）")

st.write("---")

if uploaded and tax_area and owner and period:
    if st.button("生成对账表并下载"):
        df = pd.read_excel(uploaded)

        result = generate_reconcile_table(df, tax_area, owner, period)

        # 导出 Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            result.to_excel(writer, index=False, sheet_name="对账表")
            # 取出 workbook & worksheet
            workbook = writer.book
            sheet = writer.sheets["对账表"]
            # ===== 设置列宽 =====
            column_widths = {
                "A": 4,  # 序号
                "B": 16,  # 税收地
                "C": 28,  # 企业名称
                "D": 6,  # 服务费率
                "E": 6,  # 费率
                "F": 10,  # 对账周期
                "G": 8,  # 负责人
                "H": 14,  # 支付宝实发金额
                "I": 14,  # 支付宝服务费
                "J": 12,  # 网商银行实发金额
                "K": 14,  # 网商银行服务费
                "L": 12,  # 合计金额
                "M": 18,  # 开票信息
                "N": 12,  # 开票时间
                "O": 12,  # 开票金额
                "P": 12,  # 应付服务费
                "Q": 8,  # 是否付服务费
                "R": 12,  # 实际应付服务费
                "S": 12,  # 返税金额
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

                # ===== 设置表头样式：加粗、蓝底、居中 =====
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # 正式蓝
            header_font = Font(bold=True)

            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # ===== 设置金额列右对齐 =====
            money_columns = ["H", "I", "J", "K", "L", "O", "P", "R", "S"]

            for col_letter in money_columns:
                for row in range(2, sheet.max_row + 1):  # 数据行从第2行开始
                    cell = sheet[f"{col_letter}{row}"]
                    cell.alignment = Alignment(horizontal="right")

            # ===== 所有单元格统一垂直居中 =====
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    if cell.alignment.horizontal is None:  # 避免覆盖金额右对齐
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

            # ===== 设置行高 =====
            for row in range(1, sheet.max_row + 1):
                sheet.row_dimensions[row].height = 40

        excel_data = output.getvalue()
        st.success("生成成功！点击下方按钮下载：")

        st.download_button(
            label="下载对账表.xlsx",
            data=excel_data,
            file_name="对账表.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("请上传账单并填写所有表单项")
